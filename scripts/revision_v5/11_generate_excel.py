#!/usr/bin/env python
"""11_generate_excel.py (fixed) — 生成 Excel 汇总（13 工作表，修正版）。"""
import argparse, io, csv, logging, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import v5f_common as C

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", force=True)
log = logging.getLogger("v5f_excel")

SHEETS = [
    ("文件审计", "file_inventory.csv"),
    ("冻结结果复现", "freeze_reproduction_result.csv"),
    ("MMR_lambda开发集", "tables/mmr_dev_lambda_selection.csv"),
    ("MMR逐题结果", "tables/mmr_results_per_qid.csv"),
    ("MMR主结果", "tables/mmr_summary.csv"),
    ("MMR显著性", "tables/mmr_paired_bootstrap.csv"),
    ("MMR公平性检查", "tables/mmr_fairness_validation.csv"),
    ("TopK逐题结果", "tables/topk_sensitivity_per_qid.csv"),
    ("TopK主结果", "tables/topk_sensitivity_summary.csv"),
    ("TopK差值", "tables/topk_sensitivity_differences.csv"),
    ("D01-D06覆盖", "tables/topk_demand_category_coverage.csv"),
]
PAPER_TABLES = ["paper_table_mmr_main.csv", "paper_table_mmr_significance.csv",
                "paper_table_topk_sensitivity.csv", "paper_table_demand_categories.csv"]


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def write_sheet(ws, rows):
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([num(r.get(h, "")) for h in headers])
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for j, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(max(len(str(h)) + 2, 10), 60)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def load_csv(path):
    if not path.exists():
        return []
    with io.open(path, "r", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def fmt_p(p):
    p = float(p)
    return "<0.0001" if p < 0.0001 else f"{p:.4f}"


def write_paper_tables(out):
    """论文紧凑表（round 4；百分比 1 位；无星号无加粗）。"""
    # 表1 MMR 主结果
    summ = {(r["split"], r["method"]): r for r in load_csv(out / "tables/mmr_summary.csv")}
    rows1 = []
    for m in ["B0", "MMR", "D3_TFIDF"]:
        def dcnd(sp):
            r = summ.get((sp, m))
            return f"{float(r['demand_cov_5']):.4f}/{float(r['ndcg_10']):.4f}" if r else "-"
        red = summ.get(("formal_train", m))
        rows1.append({"method": m, "formal_DC_NDCG": dcnd("formal_train"),
                      "internal_DC_NDCG": dcnd("internal_blind"),
                      "external_DC_NDCG": dcnd("cmedqa2_external"),
                      "redundancy": f"{float(red['redundancy']):.4f}" if red else "-"})
    C.write_csv(out / "tables/paper_table_mmr_main.csv", rows1)
    # 表2 显著性
    sig = load_csv(out / "tables/mmr_paired_bootstrap.csv")
    rows2 = []
    for r in sig:
        if r["family_id"] not in ("PRIMARY_DC", "PRIMARY_NDCG", "DIAGNOSTIC_DC", "DIAGNOSTIC_NDCG"):
            continue
        rows2.append({"split": r["split"], "comparison": f"{r['method_a']} vs {r['method_b']}",
                      "metric": "DemandCov@5" if r["metric"] == "demand_cov" else "NDCG@10",
                      "mean_difference": f"{float(r['mean_difference']):.4f}",
                      "ci": f"[{float(r['ci_low']):.4f}, {float(r['ci_high']):.4f}]",
                      "p_raw": fmt_p(r["p_raw"]), "p_holm": fmt_p(r["p_holm"])})
    C.write_csv(out / "tables/paper_table_mmr_significance.csv", rows2)
    # 表3 Top-K（DemandCov@K / NDCG@K）
    tks = {(r["split"], r["method"], int(r["k"])): r for r in load_csv(out / "tables/topk_sensitivity_summary.csv")}
    rows3 = []
    for sp in ["formal_train", "internal_blind", "cmedqa2_external"]:
        for m in ["B0", "D0", "D3_TFIDF", "MMR"]:
            def kcell(k):
                r = tks.get((sp, m, k))
                return f"{float(r['demand_cov_at_k']):.4f}/{float(r['ndcg_at_k']):.4f}" if r else "-"
            rows3.append({"method": f"{sp}|{m}", "K3_DC_NDCG": kcell(3), "K5_DC_NDCG": kcell(5), "K7_DC_NDCG": kcell(7)})
    C.write_csv(out / "tables/paper_table_topk_sensitivity.csv", rows3)
    # 表4 需求类别
    cat = {(r["split"], r["method"], int(r["k"])): r for r in load_csv(out / "tables/topk_demand_category_coverage.csv")}
    rows4 = []
    for sp in ["formal_train", "internal_blind", "cmedqa2_external"]:
        for m in ["B0", "MMR", "D3_TFIDF"]:
            for k in [3, 5, 7]:
                row = {"method": f"{sp}|{m}|K{k}"}
                for c in C.L1:
                    v = cat.get((sp, m, k), {}).get(c, "")
                    row[c] = f"{float(v)*100:.1f}%" if v != "" else "-"
                rows4.append(row)
    C.write_csv(out / "tables/paper_table_demand_categories.csv", rows4)
    log.info("paper tables written")


def main():
    ap = argparse.ArgumentParser(description="生成 Excel 汇总（修正版）")
    ap.add_argument("--out", default=str(C.V5))
    args = ap.parse_args()
    out = Path(args.out)
    write_paper_tables(out)

    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rel in SHEETS:
        ws = wb.create_sheet(sheet_name[:31])
        rows = load_csv(out / rel)
        write_sheet(ws, rows)
        log.info("sheet %s: %d rows", sheet_name, len(rows))

    ws = wb.create_sheet("论文紧凑表")
    first = True
    for pt in PAPER_TABLES:
        rows = load_csv(out / "tables" / pt)
        if not rows:
            continue
        if not first:
            ws.append([])
        ws.append([pt])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        write_sheet(ws, rows)
        ws.append([])
        first = False
    if ws.max_row > 1:
        ws.freeze_panes = "A2"

    ws = wb.create_sheet("复现参数")
    params = [
        ("项目", "MedSAGE-RAG 无泄漏评测 + MMR/Top-K 敏感性（fixed 版）"),
        ("修复项1", "配对 Bootstrap：diff=a-b，同一组 qid 索引 idx 重采样 mean(diff[idx])"),
        ("修复项2", "gold_ev_demands：15 类 state -> 6 类 demand（ONT[s] 映射）"),
        ("修复项3", "MMR 显式 tie-break：MMR score desc -> reranker desc -> doc_id asc"),
        ("修复项4", "全精度 float64；逐 qid CSV ≥8 位小数；论文表最后 4 位"),
        ("Bootstrap 次数", "10000"), ("Bootstrap seed", "42"), ("CI", "percentile 95%"),
        ("MMR 名称", "MMR-TFIDF（基于 TF-IDF 文档相似度的 MMR 代理基线）"),
        ("Q1/E2 预测", "复用 revision_v5_mmr_topk/cache（不重新训练）"),
        ("λ 选择划分", "formal_dev"), ("冻结 λ", "见 frozen_lambda.json"),
        ("Top-K 口径", "跨 K 以 DemandCov@K、NDCG@K 为主；official NDCG@10 仅兼容 K=5"),
        ("显著性家族", "PRIMARY_DC/PRIMARY_NDCG/DIAGNOSTIC_DC/DIAGNOSTIC_NDCG/SUPPLEMENTARY"),
        ("多重比较", "每个 family 内 Holm-Bonferroni"),
    ]
    ws.append(["参数", "值"])
    for k, v in params:
        ws.append([k, v])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{ws.max_row}"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    for cell in ws[1]:
        cell.font = Font(bold=True)

    xlsx = out / "MMR与TopK敏感性实验汇总_修正版.xlsx"
    wb.save(xlsx)
    log.info("xlsx saved: %s", xlsx)
    print("DONE excel", xlsx.name)


if __name__ == "__main__":
    main()
